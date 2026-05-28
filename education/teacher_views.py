from io import BytesIO
from urllib.parse import quote

from django.contrib import messages
from django.db.models import Count, Max, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .forms import AnswerOptionFormSet, QuestionForm, TeacherTestForm
from .models import AcademicGroup, ActivityLog, Question, StudentAnswer, Test, TestAttempt, User
from .utils import forbidden_response, log_activity, role_required


# кабинет преподавателя собирает его тесты и краткую статистику попыток
@role_required(User.Role.TEACHER)
def teacher_dashboard(request):
    tests = list(
        request.user.created_tests.select_related("discipline")
        .annotate(
            attempts_total=Count("attempts", filter=Q(attempts__is_finished=True), distinct=True),
        )
        .all()
    )
    assigned_disciplines = request.user.disciplines_taught.prefetch_related("groups")
    return render(
        request,
        "education/teacher_dashboard.html",
        {
            "tests": tests,
            "assigned_disciplines": assigned_disciplines,
            "summary": {
                "tests_total": len(tests),
                "attempts_total": sum(test.attempts_total or 0 for test in tests),
            },
        },
    )


# создание и редактирование тестов ограничено текущим преподавателем
@role_required(User.Role.TEACHER)
def test_create(request):
    if request.method == "POST":
        form = TeacherTestForm(request.POST, teacher=request.user)
        if form.is_valid():
            test = form.save(commit=False)
            test.author = request.user
            test.save()
            form.save_m2m()
            log_activity(
                request,
                request.user,
                ActivityLog.ActionType.TEST,
                f"Создан тест «{test.title}»",
            )
            messages.success(request, "Тест создан. Теперь добавьте вопросы.")
            return redirect("education:test_preview", pk=test.pk)
    else:
        form = TeacherTestForm(teacher=request.user)

    return render(
        request,
        "education/form_page.html",
        {
            "title": "Новый тест",
            "subtitle": "Задайте параметры тестирования и добавьте вопросы.",
            "form": form,
            "submit_label": "Сохранить",
            "cancel_url": "education:teacher_dashboard",
            "uniform_field_sizes": True,
            "single_column_form": True,
        },
    )


# изменение параметров теста не меняет автора и список уже созданных вопросов
@role_required(User.Role.TEACHER)
def test_update(request, pk):
    test = get_object_or_404(Test, pk=pk, author=request.user)
    if request.method == "POST":
        form = TeacherTestForm(
            request.POST,
            instance=test,
            teacher=request.user,
        )
        if form.is_valid():
            form.save()
            log_activity(
                request,
                request.user,
                ActivityLog.ActionType.TEST,
                f"Обновлён тест «{test.title}»",
            )
            messages.success(request, "Изменения сохранены.")
            return redirect("education:test_preview", pk=test.pk)
    else:
        form = TeacherTestForm(instance=test, teacher=request.user)

    return render(
        request,
        "education/form_page.html",
        {
            "title": "Редактирование теста",
            "subtitle": "Измените параметры попыток и доступности.",
            "form": form,
            "submit_label": "Сохранить",
            "cancel_url": "education:test_preview",
            "cancel_kwargs": {"pk": test.pk},
        },
    )


# удаление теста доступно только автору и выполняется через POST
@role_required(User.Role.TEACHER)
@require_POST
def test_delete(request, pk):
    test = get_object_or_404(Test, pk=pk, author=request.user)
    title = test.title
    test.delete()
    log_activity(request, request.user, ActivityLog.ActionType.TEST, f"Удалён тест «{title}»")
    messages.success(request, "Тест удалён.")
    return redirect("education:teacher_dashboard")


# предпросмотр показывает структуру теста перед прохождением студентами
@role_required(User.Role.TEACHER)
def test_preview(request, pk):
    test = get_object_or_404(
        Test.objects.select_related("discipline", "author").prefetch_related("questions__options", "groups"),
        pk=pk,
        author=request.user,
    )
    return render(request, "education/test_preview.html", {"test": test})


# вопрос сохраняется вместе с набором вариантов ответа через inline-formset
@role_required(User.Role.TEACHER)
def question_create(request, test_pk):
    test = get_object_or_404(Test, pk=test_pk, author=request.user)
    if request.method == "POST":
        form = QuestionForm(request.POST, request.FILES)
        formset = AnswerOptionFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            question = form.save(commit=False)
            question.test = test
            question.order = (test.questions.aggregate(max_order=Max("order")).get("max_order") or 0) + 1
            question.save()
            formset.instance = question
            formset.save()
            log_activity(
                request,
                request.user,
                ActivityLog.ActionType.TEST,
                f"Добавлен вопрос в тест «{test.title}»",
            )
            messages.success(request, "Вопрос добавлен.")
            return redirect("education:test_preview", pk=test.pk)
    else:
        form = QuestionForm()
        formset = AnswerOptionFormSet()

    return render(
        request,
        "education/form_page.html",
        {
            "title": "Добавление вопроса",
            "subtitle": "Укажите текст вопроса, изображение и варианты ответов.",
            "form": form,
            "formset": formset,
            "submit_label": "Сохранить",
            "cancel_url": "education:test_preview",
            "cancel_kwargs": {"pk": test.pk},
        },
    )


# редактирование вопроса работает только внутри теста текущего преподавателя
@role_required(User.Role.TEACHER)
def question_update(request, test_pk, question_pk):
    test = get_object_or_404(Test, pk=test_pk, author=request.user)
    question = get_object_or_404(Question, pk=question_pk, test=test)
    if request.method == "POST":
        form = QuestionForm(request.POST, request.FILES, instance=question)
        formset = AnswerOptionFormSet(request.POST, instance=question)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            log_activity(
                request,
                request.user,
                ActivityLog.ActionType.TEST,
                f"Изменён вопрос в тесте «{test.title}»",
            )
            messages.success(request, "Вопрос обновлён.")
            return redirect("education:test_preview", pk=test.pk)
    else:
        form = QuestionForm(instance=question)
        formset = AnswerOptionFormSet(instance=question)

    return render(
        request,
        "education/form_page.html",
        {
            "title": "Редактирование вопроса",
            "subtitle": "Измените формулировку, изображение и правильный вариант ответа.",
            "form": form,
            "formset": formset,
            "submit_label": "Сохранить изменения",
            "cancel_url": "education:test_preview",
            "cancel_kwargs": {"pk": test.pk},
            "file_delete_field": "image",
            "file_delete_label": "Удалить изображение",
            "file_delete_url": "education:question_image_delete",
            "file_delete_kwargs": {"test_pk": test.pk, "question_pk": question.pk},
        },
    )


# отдельное удаление картинки не трогает текст вопроса и варианты ответа
@role_required(User.Role.TEACHER)
@require_POST
def question_image_delete(request, test_pk, question_pk):
    test = get_object_or_404(Test, pk=test_pk, author=request.user)
    question = get_object_or_404(Question, pk=question_pk, test=test)
    if question.image:
        question.image.delete(save=False)
        question.image = None
        question.save(update_fields=["image"])
        log_activity(
            request,
            request.user,
            ActivityLog.ActionType.TEST,
            f"Удалено изображение вопроса в тесте «{test.title}»",
        )
        messages.success(request, "Изображение вопроса удалено.")
    else:
        messages.info(request, "У вопроса нет изображения для удаления.")
    return redirect("education:question_update", test_pk=test.pk, question_pk=question.pk)


# после удаления вопроса порядок оставшихся вопросов нормализуется
@role_required(User.Role.TEACHER)
@require_POST
def question_delete(request, test_pk, question_pk):
    test = get_object_or_404(Test, pk=test_pk, author=request.user)
    question = get_object_or_404(Question, pk=question_pk, test=test)
    question.delete()
    test.normalize_question_order()
    log_activity(
        request,
        request.user,
        ActivityLog.ActionType.TEST,
        f"Удалён вопрос из теста «{test.title}»",
    )
    messages.success(request, "Вопрос удалён.")
    return redirect("education:test_preview", pk=test.pk)


# отчет группирует завершенные попытки и считает процент правильных ответов по вопросам
@role_required(User.Role.TEACHER, User.Role.ADMINISTRATOR)
def report_detail(request, pk):
    test = get_object_or_404(Test.objects.select_related("discipline", "author"), pk=pk)
    if request.user.is_teacher and test.author != request.user:
        return forbidden_response(request, "У вас нет доступа к отчёту по этому тесту.")

    attempts = (
        test.attempts.filter(is_finished=True)
        .select_related("student", "student__academic_group")
        .prefetch_related("answers__question", "answers__selected_option")
        .order_by(
            "student__academic_group__name",
            "student__last_name",
            "student__first_name",
            "student__middle_name",
            "-completed_at",
            "-started_at",
        )
    )
    attempts_list = list(attempts)
    attempt_groups = []
    attempt_groups_by_key = {}
    for attempt in attempts_list:
        group = attempt.student.academic_group
        group_key = group.pk if group else None
        if group_key not in attempt_groups_by_key:
            group_data = {
                "group": group,
                "group_name": group.name if group else "Без группы",
                "attempts": [],
            }
            attempt_groups_by_key[group_key] = group_data
            attempt_groups.append(group_data)
        attempt_groups_by_key[group_key]["attempts"].append(attempt)

    for group_data in attempt_groups:
        group_data["attempts_count"] = len(group_data["attempts"])

    question_stats = []
    for question in test.questions.all():
        total_answers = StudentAnswer.objects.filter(
            question=question,
            attempt__test=test,
            attempt__is_finished=True,
        ).exclude(selected_option__isnull=True)
        answers_count = total_answers.count()
        correct_count = total_answers.filter(is_correct=True).count()
        question_stats.append(
            {
                "question": question,
                "answers_count": answers_count,
                "correct_rate": int(correct_count / answers_count * 100) if answers_count else 0,
            }
        )

    return render(
        request,
        "education/report_detail.html",
        {
            "test": test,
            "attempt_groups": attempt_groups,
            "question_stats": question_stats,
            "summary": {
                "attempts": len(attempts_list),
            },
        },
    )


# экспорт формирует Excel-листы по группам и дисциплинам с лучшими результатами студентов
@role_required(User.Role.TEACHER, User.Role.ADMINISTRATOR)
def test_results_export(request):
    if request.user.is_teacher:
        tests = request.user.created_tests.select_related("discipline").prefetch_related(
            "groups",
            "discipline__groups",
        )
    else:
        tests = Test.objects.select_related("discipline", "author").prefetch_related(
            "groups",
            "discipline__groups",
        )
    tests = list(tests.order_by("discipline__name", "title", "id"))
    test_ids = [test.pk for test in tests]
    test_group_ids = {
        test.pk: {group.pk for group in test.groups.all()}
        for test in tests
    }

    attempts = (
        TestAttempt.objects.filter(test_id__in=test_ids, is_finished=True)
        .select_related("student", "test", "test__discipline")
        .order_by("student_id", "test_id", "-score", "-completed_at", "-attempt_number")
    )
    best_attempts = {}
    for attempt in attempts:
        best_attempts.setdefault((attempt.student_id, attempt.test_id), attempt)

    workbook = Workbook()
    placeholder_sheet = workbook.active
    placeholder_sheet.title = "_"
    used_sheet_titles = set()
    header_font = Font(bold=True)

    def make_sheet_title(title):
        cleaned_title = "".join("_" if char in "[]:*?/\\" else char for char in title).strip() or "Группа"
        base_title = cleaned_title[:31]
        sheet_title = base_title
        counter = 2
        while sheet_title in used_sheet_titles:
            suffix = f" {counter}"
            sheet_title = f"{base_title[:31 - len(suffix)]}{suffix}"
            counter += 1
        used_sheet_titles.add(sheet_title)
        return sheet_title

    def add_results_sheet(group, discipline, sheet_tests, students):
        worksheet = workbook.create_sheet(make_sheet_title(group.name))
        last_column = max(len(sheet_tests) + 1, 1)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
        title_cell = worksheet["A1"]
        title_cell.value = f"Дисциплина: {discipline.name}"
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        worksheet.append(["ФИО студента", *[test.title for test in sheet_tests]])

        for student in students:
            row = [student.full_name]
            for test in sheet_tests:
                attempt = best_attempts.get((student.pk, test.pk))
                row.append(attempt.score if attempt else None)
            worksheet.append(row)

        for cell in worksheet[2]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in worksheet.iter_rows(min_row=3):
            row[0].alignment = Alignment(horizontal="left", vertical="center")
            for cell in row[1:]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        worksheet.column_dimensions["A"].width = 28
        for index, test in enumerate(sheet_tests, start=2):
            column = get_column_letter(index)
            worksheet.column_dimensions[column].width = min(max(len(test.title) + 4, 18), 60)
        worksheet.row_dimensions[1].height = 24
        worksheet.freeze_panes = "B3"

    tests_by_discipline = {}
    for test in tests:
        tests_by_discipline.setdefault(test.discipline_id, []).append(test)

    attempts_by_discipline = {}
    for attempt in attempts:
        if attempt.student.academic_group_id:
            attempts_by_discipline.setdefault(attempt.test.discipline_id, set()).add(
                attempt.student.academic_group_id
            )

    for discipline_id in sorted(
        tests_by_discipline,
        key=lambda pk: tests_by_discipline[pk][0].discipline.name,
    ):
        discipline_tests = tests_by_discipline[discipline_id]
        discipline = discipline_tests[0].discipline
        discipline_group_ids = {group.pk for group in discipline.groups.all()}
        assigned_group_ids = {
            group_id
            for test in discipline_tests
            for group_id in test_group_ids[test.pk]
        }
        group_ids = (
            discipline_group_ids
            | assigned_group_ids
            | attempts_by_discipline.get(discipline_id, set())
        )
        if not group_ids and any(not test_group_ids[test.pk] for test in discipline_tests):
            group_ids = set(
                User.objects.filter(
                    role=User.Role.STUDENT,
                    is_active=True,
                    academic_group__isnull=False,
                ).values_list("academic_group_id", flat=True)
            )

        groups = AcademicGroup.objects.filter(pk__in=group_ids).order_by("name")
        for group in groups:
            group_tests = [
                test
                for test in discipline_tests
                if not test_group_ids[test.pk] or group.pk in test_group_ids[test.pk]
            ]
            if not group_tests:
                continue
            students = User.objects.filter(
                role=User.Role.STUDENT,
                is_active=True,
                academic_group=group,
            ).order_by("last_name", "first_name", "middle_name", "username")
            add_results_sheet(group, discipline, group_tests, students)

    if len(workbook.worksheets) > 1:
        workbook.remove(placeholder_sheet)
    else:
        placeholder_sheet.title = "Результаты"
        placeholder_sheet.append(["Нет данных для выгрузки"])
        for cell in placeholder_sheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        placeholder_sheet.column_dimensions["A"].width = 34

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    export_date = timezone.localdate()
    filename = f"results_{export_date:%Y-%m-%d}.xlsx"
    quoted_filename = quote(f"Результаты {export_date:%d.%m.%Y}.xlsx")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=\"{filename}\"; filename*=UTF-8''{quoted_filename}"
    )
    log_activity(
        request,
        request.user,
        ActivityLog.ActionType.ANALYTICS,
        "Выгружены результаты тестов в Excel",
    )
    return response
